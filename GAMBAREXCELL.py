#!/usr/bin/env python
# coding: utf-8

# In[1]:


# import library
import cv2
import openpyxl
import PySimpleGUI as sg

# Create a GUI display using PySimpleGUI
sg.theme('SandyBeach')
layout = [[sg.Text('Please select an image file:'), sg.Input(key='-IN FILE-'), sg.FileBrowse()],
          [sg.Text('Please select the storage directory:'), sg.Input(key='-OUT FOLDER-'), sg.FolderBrowse()],
          [sg.Submit('Process'), sg.Cancel()]]

# Create windows
window = sg.Window('Create Pixel Value in Excel (Yasir Hasan-Indonesia)', layout)

# Opens a window and waits for input from the user
event, values = window.read()

# If the 'Cancel' button is pressed or the window is closed then the program exits
if event == sg.WINDOW_CLOSED or event == 'Cancel':
    exit()

# Read the image file and check the format
img = cv2.imread(values['-IN FILE-'])

if img is None:
    sg.popup_error('The file cannot be read or is not supported.')
    exit()


# Create a new Excel file
workbook = openpyxl.Workbook()

# Create a sheet for each RGB color and for grayscale values
sheets = {'Red': workbook.create_sheet('Red'),
          'Green': workbook.create_sheet('Green'),
          'Blue': workbook.create_sheet('Blue'),
          'Grayscale': workbook.create_sheet('Grayscale')}

# Takes the pixel value of each pixel in the image and writes it into an Excel file
for x in range(img.shape[1]):
    for y in range(img.shape[0]):
        b, g, r = img[y,x] # ambil nilai BGR
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)[y,x] # ambil nilai grayscale

        # Write the pixel value into the appropriate sheet
        sheets['Red'].cell(row=y+1, column=x+1, value=r)
        sheets['Green'].cell(row=y+1, column=x+1, value=g)
        sheets['Blue'].cell(row=y+1, column=x+1, value=b)
        sheets['Grayscale'].cell(row=y+1, column=x+1, value=gray)

# save the Excel file
workbook.save(values['-OUT FOLDER-'] + '/file_output.xlsx')

# close the window and exit the program
window.close()


# In[ ]:





# In[ ]:




